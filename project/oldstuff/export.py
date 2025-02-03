import pandas as pd
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def sanitize_sheet_name(sheet_name):
    """
    Sanitize the sheet name to ensure it is valid for Excel.
    """
    sanitized_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
    return sanitized_name[:31]  # Truncate to 31 characters

def parse_case_name(case_name):
    """
    Extract the base case name (e.g., "case_files\\Case1") and additional numerical values.
    """
    match = re.match(r'(case_files\\Case\d+)-([\d.]+)-([\d.]+)-(.+)', case_name)
    if match:
        base_name = match.group(1)  # "case_files\\CaseX"
        row_val = float(match.group(2))  # First number
        col_val = float(match.group(3))  # Second number
        return base_name, row_val, col_val
    return None, None, None

def save_results_to_excel(csv_file, excel_file, thresholds):
    # Configure logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    # Ensure thresholds is a dictionary
    if not isinstance(thresholds, dict):
        raise ValueError("Thresholds must be provided as a dictionary with case names as keys.")

    # Read the CSV file
    df = pd.read_csv(csv_file)

    # Create 'col_val' and 'row_val' columns by parsing 'case_name'
    df['base_name'], df['row_val'], df['col_val'] = zip(*df['case_name'].apply(parse_case_name))

    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    workbook = writer.book

    # Group data by base case name
    case_groups = df.groupby('base_name')

    for base_name, group in case_groups:
        if base_name not in thresholds:
            logging.warning(f"No threshold found for case: {base_name}")
            continue
        threshold = thresholds[base_name]

        # Sanitize sheet name and create sheet
        sanitized_name = sanitize_sheet_name(base_name)
        if sanitized_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sanitized_name)
        else:
            sheet = workbook[sanitized_name]

        # Log the threshold being applied
        logging.info(f"Exporting case: {base_name}, Threshold: {threshold}")

        # Write headers
        sheet.cell(row=1, column=1, value="Row\\Col")
        cols = group['col_val'].unique()
        rows = group['row_val'].unique()
        for idx, col in enumerate(cols, start=2):
            sheet.cell(row=1, column=idx, value=col)
        for idx, row in enumerate(rows, start=2):
            sheet.cell(row=idx, column=1, value=row)

        # Write data and apply conditional formatting
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for _, row in group.iterrows():
            row_idx = rows.tolist().index(row['row_val']) + 2
            col_idx = cols.tolist().index(row['col_val']) + 2
            cell = sheet.cell(row=row_idx, column=col_idx, value=row['maximum_bs_curvature'])

            # Apply formatting
            if row['maximum_bs_curvature'] < threshold['maximum_bs_curvature']:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # Save the Excel file
    writer.close()

# Example usage
if __name__ == "__main__":
    # Define thresholds as a dictionary with case names as keys
    thresholds = {
        "case_files\\Case1": {"maximum_bs_curvature": 0.055835668, "maximum_curvature": 0.055835668},
        "case_files\\Case2": {"maximum_bs_curvature": 0.051167134, "maximum_curvature": 0.051167134},
        "case_files\\Case3": {"maximum_bs_curvature": 0.046473762, "maximum_curvature": 0.046473762},
        "case_files\\Case4": {"maximum_bs_curvature": 0.041759232, "maximum_curvature": 0.041759232},
        "case_files\\Case5": {"maximum_bs_curvature": 0.034737267, "maximum_curvature": 0.034737267},
        "case_files\\Case6": {"maximum_bs_curvature": 0.032398017, "maximum_curvature": 0.032398017},
        "case_files\\Case7": {"maximum_bs_curvature": 0.030050334, "maximum_curvature": 0.030050334},
    }

    # Call the function with the correct parameters
    save_results_to_excel('bsengine-summary.csv', 'results.xlsx', thresholds)