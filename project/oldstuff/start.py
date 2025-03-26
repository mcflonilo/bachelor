import os
import subprocess
from queue import Queue
import threading
import time
import re
import csv
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook
import tkinter as tk
import logging
#from modules.maxcurve_module import DataPlotter

######################################################
# INPUT UMBILICAL DATA
length = 5.0  # length of the umbilical outside the BS SRIS
#NREL MANGLER???
EA = 1040000  # Axial stiffness [kN]
EI = 185.0    # Bending stiffness [kNm2]
GT = 143.0    # Torsional stiffness [kNm2]
m = 88.4      # Mass per unit length [kg/m]
umbilical_outer_diameter = 0.22  # outer diameter of the umbilical
#####################################################

#####################################################
# INPUT LOAD CASES

cases = ['4.0  550.0', '18.0  650.0', '23.0  750.0', '23.0  850.0', '16.0  1000.0', '10.0  1050.0', '2.0  1100.0']
#cases = ['4.0  550.0', '18.0  650.0']
# cases = ['15.0  470.0', '20.0  490.0', '25.0  700.0', '26.0  800.0', '25.0  930.0', '20.0  1020.0', '5.0  1140.0']
#cases = ['23.0  850.0']
#####################################################

#####################################################
# INPUT BS DIMENSIONS TO CALCULATE RESPONSE FOR
ODtolerance = 0.003
ODClearance = 0.010
ID = umbilical_outer_diameter+ODtolerance+(2*ODClearance) # inner diameter
#NSEG = 3 # number of segments
SL = 0.700 # segment length
#NSEL = 50 # number of elements
OD = ['1.559'] # ['1.559'] # ['1.409', '1.434', '1.459', '1.484', '1.509', '1.534', '1.559', '1.584', '1.609', '1.634', '1.659']  # ['1.534'] #, '0.250', '0.300', '0.350', '0.400', '0.450', '0.500', '0.550', '0.600', '0.650', '0.700', '0.750', '0.800', '0.850', '0.900', '0.950', '1.000', '1.050', '1.100', '1.150', '1.200']
#OD2 
CL = ['12.799'] # ['12.799'] #['11.299', '11.549', '11.799', '12.049', '12.299', '12.549', '12.799', '13.049', '13.299', '13.549', '13.799']  # ['12.594'] #, '1.75', '2.00', '2.25', '2.50', '2.75', '3.00', '3.25', '3.50', '3.75', '4.00', '4.25', '4.50', '4.75', '5.00', '5.25', '5.50', '5.75', '6.00', '6.25', '6.50', '6.75', '7.00', '7.25', '7.50', '7.75', '8.00']
#cone length
TL = 0.150 #tip length
TOD = ID+0.04 #tip outer diameter
MAT = ['NOLIN  60D_30'] #250000'] #, 'LIN  162300', 'NOLIN  60D-15deg', 'NOLIN  60D-30.8deg']
MATID = ['60D_30'] #, '162300', '60D-15deg', '60D-30.8deg']
print(ID)
thresholds = {
        "case_files\\Case1": {"maximum_bs_curvature": 0.055835668, "maximum_curvature": 0.055835668},
        "case_files\\Case2": {"maximum_bs_curvature": 0.051167134, "maximum_curvature": 0.051167134},
        "case_files\\Case3": {"maximum_bs_curvature": 0.046473762, "maximum_curvature": 0.046473762},
        "case_files\\Case4": {"maximum_bs_curvature": 0.041759232, "maximum_curvature": 0.041759232},
        "case_files\\Case5": {"maximum_bs_curvature": 0.034737267, "maximum_curvature": 0.034737267},
        "case_files\\Case6": {"maximum_bs_curvature": 0.032398017, "maximum_curvature": 0.032398017},
        "case_files\\Case7": {"maximum_bs_curvature": 0.030050334, "maximum_curvature": 0.030050334},
    }

#bslength = sl + cl + tl
#####################################################

class BSParams:
    def __init__(self, 
                 ODtolerance=0.003, 
                 ODClearance=0.010, 
                 umbilical_outer_diameter=0.22,  # Default example value
                 SL=0.700, 
                 OD=['1.559'], 
                 CL=['12.799'], 
                 TL=0.150, 
                 MAT=['NOLIN  60D_30'], 
                 MATID=['60D_30']):
        """
        Initialize BSParams with default or user-specified values.

        Parameters:
        - ODtolerance (float): Outer diameter tolerance
        - ODClearance (float): Outer diameter clearance
        - umbilical_outer_diameter (float): Umbilical's outer diameter
        - SL (float): Segment length
        - OD (list): List of outer diameters
        - CL (list): List of cone lengths
        - TL (float): Tip length
        - MAT (list): List of materials
        - MATID (list): List of material IDs
        """
        self.ODtolerance = ODtolerance
        self.ODClearance = ODClearance
        self.umbilical_outer_diameter = umbilical_outer_diameter
        self.ID = (self.umbilical_outer_diameter + self.ODtolerance + 
                   (2 * self.ODClearance))  # Inner diameter
        self.SL = SL
        self.OD = OD
        self.CL = CL
        self.TL = TL
        self.TOD = self.ID + 0.04  # Tip outer diameter
        self.MAT = MAT
        self.MATID = MATID

class RiserParams:
    def __init__(self, length, EA, EI, GT, m):
        self.length = length
        self.EA = EA
        self.EI = EI
        self.GT = GT
        self.m = m

    def __init__(self):
        self.length = 5.0
        self.EA = 1040000
        self.EI = 185.0
        self.GT = 143.0
        self.m = 88.4
        self.umbilicalOD = 0.22
def switch_frame(frame):
    frame.tkraise()


def cl_od_to_array(min_cl, max_cl, min_od, max_od, incrementSize_fieldLength, incrementSize_fieldWidth):
    cl = min_cl
    od = min_od
    cl_array = []
    od_array = []
    while cl <= max_cl:
        cl_array.append(round(cl, 3))
        cl += incrementSize_fieldLength
    while od <= max_od:
        od_array.append(round(od, 3))
        od += incrementSize_fieldWidth

    print(cl_array)
    print(od_array)

    return cl_array, od_array
    
def runBSEngine(case, case_queue):
        current_directory = os.getcwd()
        process = subprocess.Popen(f".\\bsengine -b {case}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=current_directory)
        stdout, stderr = process.communicate()
        print(f"stdout: {stdout}")
        print(f"stderr: {stderr}")
        if process.returncode != 0:
            print(f"Error encountered with case: {case}. Re-queuing the case.")
            case_queue.put(case)

def createCaseQueue():
    cases = open('bsengine-cases.txt', 'r').readlines()
    cases = [case.strip() for case in cases]
    case_queue = Queue()
    for case in cases:
        case_queue.put(case)
    return case_queue

def run_threads():
    case_queue = Queue()
    case_queue = createCaseQueue()
    print(f"Number of cases: {case_queue.qsize()}.")
    max_threads = os.cpu_count() - 1
    print(f"Max threads: {max_threads}.")
    threads = []

    while not case_queue.empty():
        while threading.active_count() > max_threads:
            time.sleep(1)
        case = case_queue.get()
        print(f"Running case: {case}.")
        thread = threading.Thread(target=runBSEngine, args=(case,case_queue))
        time.sleep(1)
        thread.start()
        threads.append(thread)
        print(f"Active threads: {threading.active_count()}.")

    for thread in threads:
        thread.join()
    print("All cases have been processed.")

def createCSVResultFile(case_list_file, summary_file_name):
    def extract_key_results(case_file):
        try:
            with open(case_file, 'r') as res_file:
                res_lines = res_file.readlines()

            keyres1 = None
            keyres2 = None

            # Search for key results in the log file
            for line in res_lines:
                if re.search(r'Maximum BS curvature', line):
                    keyres1 = line.strip().split(':')[-1].strip()
                if re.search(r'Maximum curvature', line):
                    keyres2 = line.strip().split(':')[-1].strip()

            if keyres1 and keyres2:
                return {
                    "case_name": case_file.rstrip('.log'),
                    "maximum_bs_curvature": keyres1,
                    "maximum_curvature": keyres2
                }
            else:
                print(f"Key results not found in {case_file}")
                return None

        except FileNotFoundError:
            print(f"File not found: {case_file}")
            return None
        except Exception as e:
            print(f"Error processing {case_file}: {e}")
            return None

    try:
        with open(case_list_file, 'r') as f:
            cases = f.readlines()

        results = []
        for case in cases:
            case_file = case.strip().replace('.inp', '.log')
            result = extract_key_results(case_file)
            if result:
                results.append(result)

        with open(summary_file_name, 'w', newline='') as summary_file:
            writer = csv.DictWriter(summary_file, fieldnames=["case_name", "maximum_bs_curvature", "maximum_curvature"])
            writer.writeheader()
            writer.writerows(results)

    except FileNotFoundError:
        print(f"File not found: {case_list_file}")
    except Exception as e:
        print(f"Error processing {case_list_file}: {e}")



def save_results_to_excel(csv_file, excel_file, thresholds):
    def sanitize_sheet_name(sheet_name):
        """
        Sanitize the sheet name to ensure it is valid for Excel.
        """
        sanitized_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
        return sanitized_name[:31]  # Truncate to 31 characters

    def parse_case_name(case_name):
        """
        Extract the base case name (e.g., "case_files\\Case1") and additional numerical values.
        """
        match = re.match(r'(case_files\\Case\d+)-([\d.]+)-([\d.]+)-(.+)', case_name)
        if match:
            base_name = match.group(1)  # "case_files\\CaseX"
            row_val = float(match.group(2))  # First number
            col_val = float(match.group(3))  # Second number
            return base_name, row_val, col_val
        return None, None, None
    # Configure logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    # Ensure thresholds is a dictionary
    if not isinstance(thresholds, dict):
        raise ValueError("Thresholds must be provided as a dictionary with case names as keys.")

    # Read the CSV file
    df = pd.read_csv(csv_file)

    # Create 'col_val' and 'row_val' columns by parsing 'case_name'
    df['base_name'], df['row_val'], df['col_val'] = zip(*df['case_name'].apply(parse_case_name))

    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    workbook = writer.book

    # Group data by base case name
    case_groups = df.groupby('base_name')

    for base_name, group in case_groups:
        if base_name not in thresholds:
            logging.warning(f"No threshold found for case: {base_name}")
            continue
        threshold = thresholds[base_name]

        # Sanitize sheet name and create sheet
        sanitized_name = sanitize_sheet_name(base_name)
        if sanitized_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sanitized_name)
        else:
            sheet = workbook[sanitized_name]

        # Log the threshold being applied
        logging.info(f"Exporting case: {base_name}, Threshold: {threshold}")

        # Write headers
        sheet.cell(row=1, column=1, value="Row\\Col")
        cols = group['col_val'].unique()
        rows = group['row_val'].unique()
        for idx, col in enumerate(cols, start=2):
            sheet.cell(row=1, column=idx, value=col)
        for idx, row in enumerate(rows, start=2):
            sheet.cell(row=idx, column=1, value=row)

        # Write data and apply conditional formatting
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for _, row in group.iterrows():
            row_idx = rows.tolist().index(row['row_val']) + 2
            col_idx = cols.tolist().index(row['col_val']) + 2
            cell = sheet.cell(row=row_idx, column=col_idx, value=row['maximum_bs_curvature'])

            # Apply formatting
            if row['maximum_bs_curvature'] < threshold['maximum_bs_curvature']:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # Save the Excel file
    writer.close()

def generate_case_files(length, EA, EI, GT, m, cases, ID, SL, OD, CL, TL, TOD, MAT, MATID):
    inp1 = open('bsengine-cases.txt', 'w')

    x = 0 # Counter
    y = 0
    

    for i in cases:
        x = 0
        y = y + 1
        for j in OD:
            x = 0
            for k in CL:
                x = 0
                for l in MAT:
                    x = 0
                    q = MATID[x]
                    x = x + 1
                    # Create a directory for the case files if it doesn't exist
                    if not os.path.exists('case_files'):
                        os.makedirs('case_files')

                    inp = open(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp"), 'w')
                    inp1.write(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp") + '\n')

                    inp.write('BEND STIFFENER DATA'+'\n')
                    inp.write("' ID   NSEG"+'\n')
                    inp.write("' inner diameter      Number of linear segments"+'\n')
                    inp.write(str(ID) + "   3"+'\n')
                    inp.write("' LENGTH   NEL   OD1    OD2  DENSITY LIN/NOLIN        EMOD/MAT-ID"+'\n')
                    inp.write(str(SL) + "  50  " + str(j) + "  " + str(j) + "  2000  LIN  10000.E06"+'\n')
                    inp.write(str(k) + "  100  " + str(j) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write(str(TL) + "  50  " + str(TOD) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write("'-----------------------------------------------------------------------"+'\n')
                    inp.write("'"+'\n')
                    inp.write("RISER DATA"+'\n')
                    inp.write("'SRIS,  NEL   EI,    EA,      GT     Mass"+'\n')
                    inp.write("' (m)         kNm^2  kN              kg/m"+'\n')
                    inp.write(str(length) + "  100  " + str(EI) + "  " + str(EA) + "  " + str(GT) + "  " + str(m) +'\n')
                    inp.write("'"+'\n')
                    inp.write("' mandatory data group"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("ELEMENT PRINT"+'\n')
                    inp.write("'NSPEC"+'\n')
                    inp.write("3"+'\n')
                    inp.write("'IEL1    IEL2"+'\n')
                    inp.write("1         9"+'\n')
                    inp.write("10        30"+'\n')
                    inp.write("70        80"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE SYSTEM DATA TEST PRINT"+'\n')
                    inp.write("'IFSPRI 1/2"+'\n')
                    inp.write("1"+'\n')
                    inp.write("'2"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE ANALYSIS PARAMETERS"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'  finite element analysis parameters"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'TOLNOR  MAXIT"+'\n')
                    inp.write("1.E-07  30"+'\n')
                    inp.write("'DSINC,DSMIN,DSMAX,"+'\n')
                    inp.write("0.01  0.001  0.1"+'\n')
                    inp.write("'3.0  0.01 10."+'\n')
                    inp.write("'5.  0.1 10."+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("CURVATURE RANGE"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("'CURMAX  - Maximum curvature"+'\n')
                    inp.write("'NCURV   - Number of curvature levels"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'CURMAX (1/m),NCURV"+'\n')
                    inp.write("'0.5       30"+'\n')
                    inp.write("0.2       100"+'\n')
                    inp.write("'---------------------------------------------------"+'\n')
                    inp.write("FORCE"+'\n')
                    inp.write("'Relang  tension"+'\n')
                    inp.write("'(deg)   (kN)"+'\n')
                    inp.write(str(i) + '\n')
                    inp.write("'8.00   400.0"+'\n')
                    inp.write("'16.5   500.0"+'\n')
                    inp.write("'19.0   550.0"+'\n')
                    inp.write("'19.1   600.0"+'\n')
                    inp.write("'18.6   650.0"+'\n')
                    inp.write("'17.5   700.0"+'\n')
                    inp.write("'14.0   775.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_15"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.0 0.0"+'\n')
                    inp.write("0.005   1.40E+03"+'\n')
                    inp.write("0.010   2.57E+03"+'\n')
                    inp.write("0.015   3.61E+03"+'\n')
                    inp.write("0.020   4.55E+03"+'\n')
                    inp.write("0.025   5.36E+03"+'\n')
                    inp.write("0.030   6.03E+03"+'\n')
                    inp.write("0.035   6.59E+03"+'\n')
                    inp.write("0.040   7.02E+03"+'\n')
                    inp.write("0.045   7.37E+03"+'\n')
                    inp.write("0.050   7.67E+03"+'\n')
                    inp.write("0.055   7.92E+03"+'\n')
                    inp.write("0.060   8.13E+03"+'\n')
                    inp.write("0.065   8.31E+03"+'\n')
                    inp.write("0.070   8.47E+03"+'\n')
                    inp.write("0.075   8.61E+03"+'\n')
                    inp.write("0.080   8.74E+03"+'\n')
                    inp.write("0.085   8.86E+03"+'\n')
                    inp.write("0.090   8.96E+03"+'\n')
                    inp.write("0.095   9.06E+03"+'\n')
                    inp.write("0.100   9.10E+03"+'\n')
                    inp.write("'"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_30"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.000   0.0"+'\n')
                    inp.write("0.005   1100.0"+'\n')
                    inp.write("0.010   2060.0"+'\n')
                    inp.write("0.015   2910.0"+'\n')
                    inp.write("0.020   3690.0"+'\n')
                    inp.write("0.025   4370.0"+'\n')
                    inp.write("0.030   4950.0"+'\n')
                    inp.write("0.035   5420.0"+'\n')
                    inp.write("0.040   5810.0"+'\n')
                    inp.write("0.045   6120.0"+'\n')
                    inp.write("0.050   6400.0"+'\n')
                    inp.write("0.055   6640.0"+'\n')
                    inp.write("0.060   6840.0"+'\n')
                    inp.write("0.065   7030.0"+'\n')
                    inp.write("0.070   7180.0"+'\n')
                    inp.write("0.075   7330.0"+'\n')
                    inp.write("0.080   7470.0"+'\n')
                    inp.write("0.085   7590.0"+'\n')
                    inp.write("0.090   7710.0"+'\n')
                    inp.write("0.095   7810.0"+'\n')
                    inp.write("0.100   7920.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'EXPORT MATERIAL DATA"+'\n')
                    inp.write("'--------------------------------------------------"+'\n')
                    inp.write("' IMEX   = 1 : tabular  =2 riflex format"+'\n')
                    inp.write("'  1"+'\n')
                    inp.write("'---------------------------------------------"+'\n')
                    inp.write("end"+'\n')
                    inp.write("'mandatory data group"+'\n')
                    inp.write("'---------------------------------------------"+'\n')

    inp.close()
    inp1.close()
       
def makeGui():
        # Create the main window
        root = tk.Tk()
        root.title("GUI with Buttons and Input Fields")

        # Create main frame
        main_frame = tk.Frame(root)
        main_frame.grid(row=0, column=0, sticky="nsew")

        # Create maxcurve frame
        maxcurve_frame = tk.Frame(root)
        maxcurve_frame.grid(row=0, column=0, sticky="nsew")

        # Create input fields and labels for umbilical parameters
        length_label = tk.Label(main_frame, text="Length:")
        length_label.grid(row=0, column=0, padx=10, pady=10)
        length_field = tk.Entry(main_frame, width=50)
        length_field.insert(0, str(length))
        length_field.grid(row=0, column=1, padx=10, pady=10)

        ea_label = tk.Label(main_frame, text="EA:")
        ea_label.grid(row=1, column=0, padx=10, pady=10)
        ea_field = tk.Entry(main_frame, width=50)
        ea_field.insert(0, str(EA))
        ea_field.grid(row=1, column=1, padx=10, pady=10)

        ei_label = tk.Label(main_frame, text="EI:")
        ei_label.grid(row=2, column=0, padx=10, pady=10)
        ei_field = tk.Entry(main_frame, width=50)
        ei_field.insert(0, str(EI))
        ei_field.grid(row=2, column=1, padx=10, pady=10)

        gt_label = tk.Label(main_frame, text="GT:")
        gt_label.grid(row=3, column=0, padx=10, pady=10)
        gt_field = tk.Entry(main_frame, width=50)
        gt_field.insert(0, str(GT))
        gt_field.grid(row=3, column=1, padx=10, pady=10)

        m_label = tk.Label(main_frame, text="Mass per unit length:")
        m_label.grid(row=4, column=0, padx=10, pady=10)
        m_field = tk.Entry(main_frame, width=50)
        m_field.insert(0, str(m))
        m_field.grid(row=4, column=1, padx=10, pady=10)

        cases_label = tk.Label(main_frame, text="Cases:")
        cases_label.grid(row=5, column=0, padx=10, pady=10)
        cases_field = tk.Entry(main_frame, width=50)
        cases_field.insert(0, ', '.join(cases))
        cases_field.grid(row=5, column=1, padx=10, pady=10)

        min_length_label = tk.Label(main_frame, text="Min Length:")
        min_length_label.grid(row=6, column=0, padx=10, pady=10)
        min_length_field = tk.Entry(main_frame, width=50)
        min_length_field.insert(0, str(11.299))
        min_length_field.grid(row=6, column=1, padx=10, pady=10)

        max_length_label = tk.Label(main_frame, text="Max Length:")
        max_length_label.grid(row=7, column=0, padx=10, pady=10)
        max_length_field = tk.Entry(main_frame, width=50)
        max_length_field.insert(0, str(13.799))
        max_length_field.grid(row=7, column=1, padx=10, pady=10)

        min_diameter_label = tk.Label(main_frame, text="Min Diameter:")
        min_diameter_label.grid(row=8, column=0, padx=10, pady=10)
        min_diameter_field = tk.Entry(main_frame, width=50)
        min_diameter_field.insert(0, str(1.409))
        min_diameter_field.grid(row=8, column=1, padx=10, pady=10)

        max_diameter_label = tk.Label(main_frame, text="Max Diameter:")
        max_diameter_label.grid(row=9, column=0, padx=10, pady=10)
        max_diameter_field = tk.Entry(main_frame, width=50)
        max_diameter_field.insert(0, str(1.659))
        max_diameter_field.grid(row=9, column=1, padx=10, pady=10)

        incrementSize_labelLength = tk.Label(main_frame, text="Increment size length:")
        incrementSize_labelLength.grid(row=10, column=0, padx=10, pady=10)
        incrementSize_fieldLength = tk.Entry(main_frame, width=50)
        incrementSize_fieldLength.insert(0, str(0.25))
        incrementSize_fieldLength.grid(row=10, column=1, padx=10, pady=10)

        incrementSize_labelWidth = tk.Label(main_frame, text="Increment size width:")
        incrementSize_labelWidth.grid(row=11, column=0, padx=10, pady=10)
        incrementSize_fieldWidth = tk.Entry(main_frame, width=50)
        incrementSize_fieldWidth.insert(0, str(0.025))
        incrementSize_fieldWidth.grid(row=11, column=1, padx=10, pady=10)

        id_label = tk.Label(main_frame, text="ID:")
        id_label.grid(row=12, column=0, padx=10, pady=10)
        id_field = tk.Entry(main_frame, width=50)
        id_field.insert(0, str(ID))
        id_field.grid(row=12, column=1, padx=10, pady=10)

        sl_label = tk.Label(main_frame, text="SL:")
        sl_label.grid(row=13, column=0, padx=10, pady=10)
        sl_field = tk.Entry(main_frame, width=50)
        sl_field.insert(0, str(SL))
        sl_field.grid(row=13, column=1, padx=10, pady=10)

        od_label = tk.Label(main_frame, text="OD:")
        od_label.grid(row=14, column=0, padx=10, pady=10)
        od_field = tk.Entry(main_frame, width=50)
        od_field.insert(0, ', '.join(OD))
        od_field.grid(row=14, column=1, padx=10, pady=10)

        cl_label = tk.Label(main_frame, text="CL:")
        cl_label.grid(row=15, column=0, padx=10, pady=10)
        cl_field = tk.Entry(main_frame, width=50)
        cl_field.insert(0, ', '.join(CL))
        cl_field.grid(row=15, column=1, padx=10, pady=10)

        tl_label = tk.Label(main_frame, text="TL:")
        tl_label.grid(row=16, column=0, padx=10, pady=10)
        tl_field = tk.Entry(main_frame, width=50)
        tl_field.insert(0, str(TL))
        tl_field.grid(row=16, column=1, padx=10, pady=10)

        tod_label = tk.Label(main_frame, text="TOD:")
        tod_label.grid(row=17, column=0, padx=10, pady=10)
        tod_field = tk.Entry(main_frame, width=50)
        tod_field.insert(0, str(TOD))
        tod_field.grid(row=17, column=1, padx=10, pady=10)

        mat_label = tk.Label(main_frame, text="MAT:")
        mat_label.grid(row=18, column=0, padx=10, pady=10)
        mat_field = tk.Entry(main_frame, width=50)
        mat_field.insert(0, ', '.join(MAT))
        mat_field.grid(row=18, column=1, padx=10, pady=10)

        matid_label = tk.Label(main_frame, text="MATID:")
        matid_label.grid(row=19, column=0, padx=10, pady=10)
        matid_field = tk.Entry(main_frame, width=50)
        matid_field.insert(0, ', '.join(MATID))
        matid_field.grid(row=19, column=1, padx=10, pady=10)


        def createResults():
            createCSVResultFile('bsengine-cases.txt', 'bsengine-summary.csv')
            save_results_to_excel('bsengine-summary.csv', 'bsengine-summary.xlsx', thresholds)

        def completeAnalysis():
            CL, OD = cl_od_to_array(float(min_length_field.get()), float(max_length_field.get()), float(min_diameter_field.get()), float(max_diameter_field.get()), float(incrementSize_fieldLength.get()), float(incrementSize_fieldWidth.get()))
            generate_case_files(float(length_field.get()), float(ea_field.get()), float(ei_field.get()), float(gt_field.get()), float(m_field.get()),cases_field.get().split(', '), float(id_field.get()), float(sl_field.get()), OD,CL, float(tl_field.get()), float(tod_field.get()), mat_field.get().split(', '),matid_field.get().split(', '))
            run_threads()
            createResults()

        def generateCaseFilesProper():
            CL, OD = cl_od_to_array(float(min_length_field.get()), float(max_length_field.get()), float(min_diameter_field.get()), float(max_diameter_field.get()), float(incrementSize_fieldLength.get()), float(incrementSize_fieldWidth.get()))
            generate_case_files(float(length_field.get()), float(ea_field.get()), float(ei_field.get()), float(gt_field.get()), float(m_field.get()),cases_field.get().split(', '), float(id_field.get()), float(sl_field.get()), OD,CL, float(tl_field.get()), float(tod_field.get()), mat_field.get().split(', '),matid_field.get().split(', '))
        
        
        # Create buttons
        button1 = tk.Button(root, text="Run analysis", command=lambda: completeAnalysis())
        button1.grid(row=3, column=0, padx=10, pady=10)
        button3 = tk.Button(root, text="Check results without rerunning analysis", command=lambda: createResults()) 
        button3.grid(row=3, column=1, padx=10, pady=10)
        button4 = tk.Button(root, text="generate case files", command=lambda: generateCaseFilesProper())
        button4.grid(row=3, column=2, padx=10, pady=10)
        # Add a new button to switch to the maxcurve_frame
        button5 = tk.Button(main_frame, text="Open MaxCurve GUI", command=lambda: switch_frame(maxcurve_frame))
        button5.grid(row=3, column=3, padx=10, pady=10)
        button6 = tk.Button(maxcurve_frame, text="Return", command=lambda: run_threads())
        button6.pack(pady=10)

        #maxcurve_app = DataPlotter(maxcurve_frame)
        # Add a return button to switch back to the main_frame
        return_button = tk.Button(maxcurve_frame, text="Return", command=lambda: switch_frame(main_frame))
        return_button.pack(pady=10)

        # Raise the main_frame initially
        main_frame.tkraise()

        return root

def main():
    #BSParams1 = BSParams(1, 0.5, 0.2, 0.1, 0.05, 0.01, [60, 60], ['60D_15', '60D_30'])
    #generate_case_files(RiserParams1.length, RiserParams1.EA, RiserParams1.EI, RiserParams1.GT, RiserParams1.m, [8, 16.5, 19, 19.1, 18.6, 17.5, 14], BSParams1.ID, BSParams1.SL, BSParams1.OD, BSParams1.CL, BSParams1.TL, BSParams1.TOD, BSParams1.MAT, BSParams1.MATID)
    #run_threads() #running the threads will create a lot of log files, which will be used to create the csv file.
    #createCSVResultFile('bsengine-cases.txt', 'bsengine-summary.csv') #creates the csv file that is used to create the excel file.
    #save_results_to_excel('bsengine-summary.csv', 'bsengine-summary.xlsx', 0.0417592324286036) #creates the excel file with conditional formatting.
    root = makeGui()
    root.mainloop()

main()

#kan forbedres ved å finne en annen måte å starte tråder, fungerer for øyeblikket.