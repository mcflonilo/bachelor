import tkinter as tk
from tkinter import messagebox
import subprocess
import os
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

######################################################
# INPUT UMBILICAL DATA
length = 5.0  # length of the umbilical outside the BS SRIS
#NREL MANGLER???
EA = 1040000  # Axial stiffness [kN]
EI = 185.0    # Bending stiffness [kNm2]
GT = 143.0    # Torsional stiffness [kNm2]
m = 88.4      # Mass per unit length [kg/m]
umbilical_outer_diameter = 0.22  # outer diameter of the umbilical
#####################################################

#####################################################
# INPUT LOAD CASES

#cases = ['4.0  550.0', '18.0  650.0', '23.0  750.0', '23.0  850.0', '16.0  1000.0', '10.0  1050.0', '2.0  1100.0']
#cases = ['4.0  550.0', '18.0  650.0']
# cases = ['15.0  470.0', '20.0  490.0', '25.0  700.0', '26.0  800.0', '25.0  930.0', '20.0  1020.0', '5.0  1140.0']
cases = ['23.0  850.0']
#####################################################

#####################################################
# INPUT BS DIMENSIONS TO CALCULATE RESPONSE FOR
ODtolerance = 0.003
ODClearance = 0.010
ID = umbilical_outer_diameter+ODtolerance+(2*ODClearance) # inner diameter
#NSEG = 3 # number of segments
SL = 0.700 # segment length
#NSEL = 50 # number of elements
OD = ['1.559'] # ['1.559'] # ['1.409', '1.434', '1.459', '1.484', '1.509', '1.534', '1.559', '1.584', '1.609', '1.634', '1.659']  # ['1.534'] #, '0.250', '0.300', '0.350', '0.400', '0.450', '0.500', '0.550', '0.600', '0.650', '0.700', '0.750', '0.800', '0.850', '0.900', '0.950', '1.000', '1.050', '1.100', '1.150', '1.200']
#OD2 
CL = ['12.799'] # ['12.799'] #['11.299', '11.549', '11.799', '12.049', '12.299', '12.549', '12.799', '13.049', '13.299', '13.549', '13.799']  # ['12.594'] #, '1.75', '2.00', '2.25', '2.50', '2.75', '3.00', '3.25', '3.50', '3.75', '4.00', '4.25', '4.50', '4.75', '5.00', '5.25', '5.50', '5.75', '6.00', '6.25', '6.50', '6.75', '7.00', '7.25', '7.50', '7.75', '8.00']
#cone length
TL = 0.150 #tip length
TOD = ID+0.04 #tip outer diameter
MAT = ['NOLIN  60D_30'] #250000'] #, 'LIN  162300', 'NOLIN  60D-15deg', 'NOLIN  60D-30.8deg']
MATID = ['60D_30'] #, '162300', '60D-15deg', '60D-30.8deg']

#bslength = sl + cl + tl
#####################################################

def generate_case_files(length, EA, EI, GT, m, cases, ID, SL, OD, CL, TL, TOD, MAT, MATID):
    inp1 = open('bsengine-cases.txt', 'w')

    x = 0 # Counter
    y = 0
    

    for i in cases:
        x = 0
        y = y + 1
        for j in OD:
            x = 0
            for k in CL:
                x = 0
                for l in MAT:
                    x = 0
                    q = MATID[x]
                    x = x + 1
                    # Create a directory for the case files if it doesn't exist
                    if not os.path.exists('case_files'):
                        os.makedirs('case_files')

                    inp = open(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp"), 'w')
                    inp1.write(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp") + '\n')

                    inp.write('BEND STIFFENER DATA'+'\n')
                    inp.write("' ID   NSEG"+'\n')
                    inp.write("' inner diameter      Number of linear segments"+'\n')
                    inp.write(str(ID) + "   3"+'\n')
                    inp.write("' LENGTH   NEL   OD1    OD2  DENSITY LIN/NOLIN        EMOD/MAT-ID"+'\n')
                    inp.write(str(SL) + "  50  " + str(j) + "  " + str(j) + "  2000  LIN  10000.E06"+'\n')
                    inp.write(str(k) + "  100  " + str(j) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write(str(TL) + "  50  " + str(TOD) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write("'-----------------------------------------------------------------------"+'\n')
                    inp.write("'"+'\n')
                    inp.write("RISER DATA"+'\n')
                    inp.write("'SRIS,  NEL   EI,    EA,      GT     Mass"+'\n')
                    inp.write("' (m)         kNm^2  kN              kg/m"+'\n')
                    inp.write(str(length) + "  100  " + str(EI) + "  " + str(EA) + "  " + str(GT) + "  " + str(m) +'\n')
                    inp.write("'"+'\n')
                    inp.write("' mandatory data group"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("ELEMENT PRINT"+'\n')
                    inp.write("'NSPEC"+'\n')
                    inp.write("3"+'\n')
                    inp.write("'IEL1    IEL2"+'\n')
                    inp.write("1         9"+'\n')
                    inp.write("10        30"+'\n')
                    inp.write("70        80"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE SYSTEM DATA TEST PRINT"+'\n')
                    inp.write("'IFSPRI 1/2"+'\n')
                    inp.write("1"+'\n')
                    inp.write("'2"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE ANALYSIS PARAMETERS"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'  finite element analysis parameters"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'TOLNOR  MAXIT"+'\n')
                    inp.write("1.E-07  30"+'\n')
                    inp.write("'DSINC,DSMIN,DSMAX,"+'\n')
                    inp.write("0.01  0.001  0.1"+'\n')
                    inp.write("'3.0  0.01 10."+'\n')
                    inp.write("'5.  0.1 10."+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("CURVATURE RANGE"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("'CURMAX  - Maximum curvature"+'\n')
                    inp.write("'NCURV   - Number of curvature levels"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'CURMAX (1/m),NCURV"+'\n')
                    inp.write("'0.5       30"+'\n')
                    inp.write("0.2       100"+'\n')
                    inp.write("'---------------------------------------------------"+'\n')
                    inp.write("FORCE"+'\n')
                    inp.write("'Relang  tension"+'\n')
                    inp.write("'(deg)   (kN)"+'\n')
                    inp.write(str(i) + '\n')
                    inp.write("'8.00   400.0"+'\n')
                    inp.write("'16.5   500.0"+'\n')
                    inp.write("'19.0   550.0"+'\n')
                    inp.write("'19.1   600.0"+'\n')
                    inp.write("'18.6   650.0"+'\n')
                    inp.write("'17.5   700.0"+'\n')
                    inp.write("'14.0   775.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_15"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.0 0.0"+'\n')
                    inp.write("0.005   1.40E+03"+'\n')
                    inp.write("0.010   2.57E+03"+'\n')
                    inp.write("0.015   3.61E+03"+'\n')
                    inp.write("0.020   4.55E+03"+'\n')
                    inp.write("0.025   5.36E+03"+'\n')
                    inp.write("0.030   6.03E+03"+'\n')
                    inp.write("0.035   6.59E+03"+'\n')
                    inp.write("0.040   7.02E+03"+'\n')
                    inp.write("0.045   7.37E+03"+'\n')
                    inp.write("0.050   7.67E+03"+'\n')
                    inp.write("0.055   7.92E+03"+'\n')
                    inp.write("0.060   8.13E+03"+'\n')
                    inp.write("0.065   8.31E+03"+'\n')
                    inp.write("0.070   8.47E+03"+'\n')
                    inp.write("0.075   8.61E+03"+'\n')
                    inp.write("0.080   8.74E+03"+'\n')
                    inp.write("0.085   8.86E+03"+'\n')
                    inp.write("0.090   8.96E+03"+'\n')
                    inp.write("0.095   9.06E+03"+'\n')
                    inp.write("0.100   9.10E+03"+'\n')
                    inp.write("'"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_30"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.000   0.0"+'\n')
                    inp.write("0.005   1100.0"+'\n')
                    inp.write("0.010   2060.0"+'\n')
                    inp.write("0.015   2910.0"+'\n')
                    inp.write("0.020   3690.0"+'\n')
                    inp.write("0.025   4370.0"+'\n')
                    inp.write("0.030   4950.0"+'\n')
                    inp.write("0.035   5420.0"+'\n')
                    inp.write("0.040   5810.0"+'\n')
                    inp.write("0.045   6120.0"+'\n')
                    inp.write("0.050   6400.0"+'\n')
                    inp.write("0.055   6640.0"+'\n')
                    inp.write("0.060   6840.0"+'\n')
                    inp.write("0.065   7030.0"+'\n')
                    inp.write("0.070   7180.0"+'\n')
                    inp.write("0.075   7330.0"+'\n')
                    inp.write("0.080   7470.0"+'\n')
                    inp.write("0.085   7590.0"+'\n')
                    inp.write("0.090   7710.0"+'\n')
                    inp.write("0.095   7810.0"+'\n')
                    inp.write("0.100   7920.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'EXPORT MATERIAL DATA"+'\n')
                    inp.write("'--------------------------------------------------"+'\n')
                    inp.write("' IMEX   = 1 : tabular  =2 riflex format"+'\n')
                    inp.write("'  1"+'\n')
                    inp.write("'---------------------------------------------"+'\n')
                    inp.write("end"+'\n')
                    inp.write("'mandatory data group"+'\n')
                    inp.write("'---------------------------------------------"+'\n')

    inp.close()
    inp1.close()

import subprocess
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from tkinter import messagebox

import os
from multiprocessing import cpu_count
from concurrent.futures import ThreadPoolExecutor, as_completed

import os
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

# Lock to ensure file access safety
file_lock = Lock()

def run_bsengine_case(case, current_directory):
    """
    Runs a single case with file lock to prevent simultaneous access to the same resource.
    """
    with file_lock:  # Lock to prevent concurrent access to files
        process = subprocess.Popen(
            f".\\bsengine -b {case}",
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            cwd=current_directory
        )
        output, error = process.communicate()
        return case, output, error

def run_exe():
    try:
        # Generate case files
        CL, OD = cl_od_to_array(
            float(min_length_field.get()), float(max_length_field.get()),
            float(min_diameter_field.get()), float(max_diameter_field.get()),
            float(incrementSize_fieldLength.get()), float(incrementSize_fieldWidth.get())
        )
        generate_case_files(
            float(length_field.get()), float(ea_field.get()), float(ei_field.get()), float(gt_field.get()), float(m_field.get()),
            cases_field.get().split(', '), float(id_field.get()), float(sl_field.get()), OD,
            CL, float(tl_field.get()), float(tod_field.get()), mat_field.get().split(', '),
            matid_field.get().split(', ')
        )
        
        # Read cases
        cases = open('bsengine-cases.txt', 'r').readlines()
        cases = [case.strip() for case in cases]
        print("Cases to run:", cases)
        
        # Get current working directory
        current_directory = os.getcwd()
        
        # Determine optimal thread count
        thread_count = get_optimal_thread_count()
        print(f"Using {thread_count} threads for execution.")
        
        # Multithreaded execution
        results = []
        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            future_to_case = {executor.submit(run_bsengine_case, case, current_directory): case for case in cases}
            for future in as_completed(future_to_case):
                case = future_to_case[future]
                try:
                    case, output, error = future.result()
                    results.append((case, output, error))
                    print(f"Case {case} completed.")
                except Exception as e:
                    print(f"Case {case} failed with error: {str(e)}")
        
        # Log the results
        for case, output, error in results:
            if error:
                print(f"Error in case {case}: {error}")
            else:
                print(f"Output for case {case}:\n{output}")

        # Process the results
        process_cases('bsengine-cases.txt', '40-results1.json')
        
        # Generate and export results
        printResults(0.0417592324286036)

    except subprocess.CalledProcessError as e:
        messagebox.showerror("Error", f"Failed to run \n{e}")
        print("Error:", e)

def get_optimal_thread_count():
    """
    Determines the optimal number of threads based on the system's CPU count.
    """
    try:
        from multiprocessing import cpu_count
        core_count = cpu_count()
        return max(1, core_count - 1)
    except Exception as e:
        print(f"Failed to determine CPU core count: {e}")
        return 4  # Fallback


def process_cases(case_list_file, summary_file_name):

    def extract_key_results(case_file):
        try:
            with open(case_file, 'r') as res_file:
                res_lines = res_file.readlines()

            keyres1 = None
            keyres2 = None

            # Search for key results in the log file
            for line in res_lines:
                if re.search(r'Maximum BS curvature', line):
                    keyres1 = line.strip().split(':')[-1].strip()
                if re.search(r'Maximum curvature', line):
                    keyres2 = line.strip().split(':')[-1].strip()

            if keyres1 and keyres2:
                return {
                    "case_name": case_file.rstrip('.log'),
                    "maximum_bs_curvature": keyres1,
                    "maximum_curvature": keyres2
                }
            else:
                print(f"Key results not found in {case_file}")
                return None

        except FileNotFoundError:
            print(f"File not found: {case_file}")
            return None
        except Exception as e:
            print(f"Error processing {case_file}: {e}")
            return None

    try:
        with open(case_list_file, 'r') as f:
            cases = f.readlines()

        results = []
        for case in cases:
            case_file = case.strip().replace('.inp', '.log')
            result = extract_key_results(case_file)
            if result:
                results.append(result)

        with open(summary_file_name, 'w') as summary_file:
            json.dump(results, summary_file, indent=4)

    except FileNotFoundError:
        print(f"File not found: {case_list_file}")
    except Exception as e:
        print(f"Error processing {case_list_file}: {e}")

def cl_od_to_array(min_cl, max_cl, min_od, max_od, incrementSize_fieldLength, incrementSize_fieldWidth):
    cl = min_cl
    od = min_od
    cl_array = []
    od_array = []
    while cl <= max_cl:
        cl_array.append(round(cl, 3))
        cl += incrementSize_fieldLength
    while od <= max_od:
        od_array.append(round(od, 3))
        od += incrementSize_fieldWidth

    print(cl_array)
    print(od_array)

    return cl_array, od_array
    
def checkResults(maxAllowdCurvature):
    with open('40-results1.json') as f:
        results = json.load(f)

    for result in results:
        if float(result['maximum_curvature']) > maxAllowdCurvature:
            print(f"Case {result['case_name']} failed with maximum curvature {result['maximum_curvature']}")
        else:
            print(f"Case {result['case_name']} passed with maximum curvature {result['maximum_curvature']}")

    save_results_to_excel(results, 'results.xlsx', maxAllowdCurvature)

def save_results_to_excel(results, filename, max_allowed_curvature):
    # Create a DataFrame from the results
    df = pd.DataFrame(results)
    
    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Convert the DataFrame to an Excel object
    df.to_excel(writer, index=False, sheet_name='Results')
    
    # Access the workbook and the sheet
    workbook = writer.book
    sheet = writer.sheets['Results']
    
    # Define the fill colors for pass and fail
    pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Iterate over the DataFrame and apply the fill colors
    for row in range(2, len(df) + 2):  # DataFrame index starts at 0, Excel row starts at 1
        max_curvature = float(sheet.cell(row=row, column=3).value)
        if max_curvature > max_allowed_curvature:
            sheet.cell(row=row, column=3).fill = fail_fill
        else:
            sheet.cell(row=row, column=3).fill = pass_fill
    
    # Save the Excel file
    writer.close()


import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def printResults(threshold):
    def generate_2d_arrays_by_case():
        # Load JSON data
        with open('40-results1.json') as f:
            results = json.load(f)

        # Initialize storage for case data
        case_data = {}

        # Process results to group by case name
        for result in results:
            case_name = result['case_name']
            parts = case_name.split('-')

            if len(parts) >= 3:
                case_prefix = parts[0]  # First part is the case prefix
                row = float(parts[1])  # Second part is row
                col = float(parts[2])  # Third part is column

                if case_prefix not in case_data:
                    case_data[case_prefix] = {"rows": set(), "cols": set(), "data": {}}

                # Add rows, columns, and data
                case_data[case_prefix]["rows"].add(row)
                case_data[case_prefix]["cols"].add(col)
                case_data[case_prefix]["data"][(row, col)] = float(result['maximum_curvature'])

        # Convert sets to sorted lists and prepare DataFrames
        for case_prefix in case_data:
            case_info = case_data[case_prefix]
            case_info["rows"] = sorted(case_info["rows"])
            case_info["cols"] = sorted(case_info["cols"])

            # Generate DataFrame
            case_info["df"] = pd.DataFrame(
                {
                    col: [case_info["data"].get((row, col), None) for col in case_info["cols"]]
                    for row in case_info["rows"]
                },
                index=case_info["rows"],
                columns=case_info["cols"],
            )

        return case_data

    def export_to_excel(df, file_name, threshold):
        # Save the DataFrame to an Excel file
        df.to_excel(file_name, sheet_name='Curvature Data', index_label='Row')

        # Apply conditional formatting
        wb = load_workbook(file_name)
        sheet = wb['Curvature Data']

        # Define color fills
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Loop through DataFrame values and apply formatting
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):  # Data starts at Excel row 2
            for col_idx, value in enumerate(row, start=2):  # Data starts at Excel column B
                if value is not None:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if value < threshold:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

        # Save the workbook
        wb.save(file_name)
        print(f"Data exported to {file_name} with conditional formatting.")

    # Generate 2D arrays for each case
    case_data = generate_2d_arrays_by_case()

    # Create output directory
    output_dir = "output_excel_files/case_files"
    os.makedirs(output_dir, exist_ok=True)

    x = 0
    # Export each case to its own Excel file
    for case_prefix, case_info in case_data.items():
        x+=1
        print(f"Processing case: {case_prefix}")
        file_name = os.path.join(output_dir, f"{x}_curvature_data.xlsx")
        export_to_excel(case_info["df"], file_name, threshold)
# Create the main window
root = tk.Tk()
root.title("GUI with Buttons and Input Fields")

# Create a frame for umbilical parameters
umbilical_frame = tk.LabelFrame(root, text="Umbilical Parameters", padx=10, pady=10)
umbilical_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

# Create input fields and labels for umbilical parameters
length_field = tk.Entry(umbilical_frame, width=50)
length_field.insert(0, str(length))
length_field.grid(row=0, column=1, padx=10, pady=10)
length_label = tk.Label(umbilical_frame, text="Length:")
length_label.grid(row=0, column=0, padx=10, pady=10)

ea_field = tk.Entry(umbilical_frame, width=50)
ea_field.insert(0, str(EA))
ea_field.grid(row=1, column=1, padx=10, pady=10)
ea_label = tk.Label(umbilical_frame, text="EA:")
ea_label.grid(row=1, column=0, padx=10, pady=10)

ei_field = tk.Entry(umbilical_frame, width=50)
ei_field.insert(0, str(EI))
ei_field.grid(row=2, column=1, padx=10, pady=10)
ei_label = tk.Label(umbilical_frame, text="EI:")
ei_label.grid(row=2, column=0, padx=10, pady=10)

gt_field = tk.Entry(umbilical_frame, width=50)
gt_field.insert(0, str(GT))
gt_field.grid(row=3, column=1, padx=10, pady=10)
gt_label = tk.Label(umbilical_frame, text="GT:")
gt_label.grid(row=3, column=0, padx=10, pady=10)

m_field = tk.Entry(umbilical_frame, width=50)
m_field.insert(0, str(m))
m_field.grid(row=4, column=1, padx=10, pady=10)
m_label = tk.Label(umbilical_frame, text="Mass per unit length:")
m_label.grid(row=4, column=0, padx=10, pady=10)

# Create a frame for load cases
cases_frame = tk.LabelFrame(root, text="Load Cases", padx=10, pady=10)
cases_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

# Create input field and label for load cases
cases_field = tk.Entry(cases_frame, width=50)
cases_field.insert(0, ', '.join(cases))
cases_field.grid(row=0, column=1, padx=10, pady=10)
cases_label = tk.Label(cases_frame, text="Cases:")
cases_label.grid(row=0, column=0, padx=10, pady=10)

# Create a frame for other parameters
other_params_frame = tk.LabelFrame(root, text="BS dimensions", padx=10, pady=10)
other_params_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

# Create input fields and labels for other parameters
# Create input fields and labels for new parameters
min_length_field = tk.Entry(other_params_frame, width=50)
min_length_field.grid(row=8, column=1, padx=10, pady=10)
min_length_field.insert(0, str(11.299))
min_length_label = tk.Label(other_params_frame, text="Min Length:")
min_length_label.grid(row=8, column=0, padx=10, pady=10)

max_length_field = tk.Entry(other_params_frame, width=50)
max_length_field.grid(row=9, column=1, padx=10, pady=10)
max_length_field.insert(0, str(13.799))
max_length_label = tk.Label(other_params_frame, text="Max Length:")
max_length_label.grid(row=9, column=0, padx=10, pady=10)

min_diameter_field = tk.Entry(other_params_frame, width=50)
min_diameter_field.grid(row=10, column=1, padx=10, pady=10)
min_diameter_field.insert(0, str(1.409))
min_diameter_label = tk.Label(other_params_frame, text="Min Diameter:")
min_diameter_label.grid(row=10, column=0, padx=10, pady=10)

max_diameter_field = tk.Entry(other_params_frame, width=50)
max_diameter_field.grid(row=11, column=1, padx=10, pady=10)
max_diameter_field.insert(0, str(1.659))
max_diameter_label = tk.Label(other_params_frame, text="Max Diameter:")
max_diameter_label.grid(row=11, column=0, padx=10, pady=10)

incrementSize_fieldLength = tk.Entry(other_params_frame, width=50)
incrementSize_fieldLength.grid(row=12, column=1, padx=10, pady=10)
incrementSize_fieldLength.insert(0, str(0.25))
incrementSize_labelLength = tk.Label(other_params_frame, text="increment size length:")
incrementSize_labelLength.grid(row=12, column=0, padx=10, pady=10)

incrementSize_fieldWidth = tk.Entry(other_params_frame, width=50)
incrementSize_fieldWidth.grid(row=13, column=1, padx=10, pady=10)
incrementSize_fieldWidth.insert(0, str(0.025))
incrementSize_labelWidth = tk.Label(other_params_frame, text="increment size width:")
incrementSize_labelWidth.grid(row=13, column=0, padx=10, pady=10)


id_field = tk.Entry(other_params_frame, width=50)
id_field.insert(0, str(ID))
id_field.grid(row=0, column=1, padx=10, pady=10)
id_label = tk.Label(other_params_frame, text="ID:")
id_label.grid(row=0, column=0, padx=10, pady=10)

sl_field = tk.Entry(other_params_frame, width=50)
sl_field.insert(0, str(SL))
sl_field.grid(row=1, column=1, padx=10, pady=10)
sl_label = tk.Label(other_params_frame, text="SL:")
sl_label.grid(row=1, column=0, padx=10, pady=10)

od_field = tk.Entry(other_params_frame, width=50)
od_field.insert(0, ', '.join(OD))
od_field.grid(row=2, column=1, padx=10, pady=10)
od_label = tk.Label(other_params_frame, text="OD:")
od_label.grid(row=2, column=0, padx=10, pady=10)

cl_field = tk.Entry(other_params_frame, width=50)
cl_field.insert(0, ', '.join(CL))
cl_field.grid(row=3, column=1, padx=10, pady=10)
cl_label = tk.Label(other_params_frame, text="CL:")
cl_label.grid(row=3, column=0, padx=10, pady=10)

tl_field = tk.Entry(other_params_frame, width=50)
tl_field.insert(0, str(TL))
tl_field.grid(row=4, column=1, padx=10, pady=10)
tl_label = tk.Label(other_params_frame, text="TL:")
tl_label.grid(row=4, column=0, padx=10, pady=10)

tod_field = tk.Entry(other_params_frame, width=50)
tod_field.insert(0, str(TOD))
tod_field.grid(row=5, column=1, padx=10, pady=10)
tod_label = tk.Label(other_params_frame, text="TOD:")
tod_label.grid(row=5, column=0, padx=10, pady=10)

mat_field = tk.Entry(other_params_frame, width=50)
mat_field.insert(0, ', '.join(MAT))
mat_field.grid(row=6, column=1, padx=10, pady=10)
mat_label = tk.Label(other_params_frame, text="MAT:")
mat_label.grid(row=6, column=0, padx=10, pady=10)

matid_field = tk.Entry(other_params_frame, width=50)
matid_field.insert(0, ', '.join(MATID))
matid_field.grid(row=7, column=1, padx=10, pady=10)
matid_label = tk.Label(other_params_frame, text="MATID:")
matid_label.grid(row=7, column=0, padx=10, pady=10)

# Create buttons
button1 = tk.Button(root, text="Run analysis", command=lambda: run_exe())
button1.grid(row=3, column=0, padx=10, pady=10)
button2 = tk.Button(root, text="Run analysis test", command=lambda: cl_od_to_array(float(min_length_field.get()), float(max_length_field.get()), float(min_diameter_field.get()), float(max_diameter_field.get()), float(incrementSize_fieldLength.get()), float(incrementSize_fieldWidth.get())))
button2.grid(row=3, column=1, padx=10, pady=10)
button3 = tk.Button(root, text="Check results", command=lambda: checkResults(0.0417592324286036))
button3.grid(row=4, column=0, padx=10, pady=10)

button4 = tk.Button(root, text="test", command=lambda: printResults(0.0417592324286036))
button4.grid(row=3, column=2, padx=10, pady=10)



# Run the main loop
root.mainloop()
# Create a button to generate case files
