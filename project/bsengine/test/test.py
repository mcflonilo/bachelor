import os
import subprocess
from queue import Queue
import threading
import time
import re
import csv
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def runBSEngine(case, case_queue):
        current_directory = os.getcwd()
        process = subprocess.Popen(f".\\bsengine -b {case}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=current_directory)
        stdout, stderr = process.communicate()
        print(f"stdout: {stdout}")
        print(f"stderr: {stderr}")
        if process.returncode != 0:
            print(f"Error encountered with case: {case}. Re-queuing the case.")
            case_queue.put(case)

def createCaseQueue():
    cases = open('testbsengine-cases.txt', 'r').readlines()
    cases = [case.strip() for case in cases]
    case_queue = Queue()
    for case in cases:
        case_queue.put(case)
    return case_queue

def run_threads():
    case_queue = Queue()
    case_queue = createCaseQueue()
    print(f"Number of cases: {case_queue.qsize()}.")
    max_threads = os.cpu_count() - 1
    print(f"Max threads: {max_threads}.")
    threads = []

    while not case_queue.empty():
        while threading.active_count() > max_threads:
            time.sleep(1)
        case = case_queue.get()
        print(f"Running case: {case}.")
        thread = threading.Thread(target=runBSEngine, args=(case,case_queue))
        time.sleep(1)
        thread.start()
        threads.append(thread)
        print(f"Active threads: {threading.active_count()}.")

    for thread in threads:
        thread.join()
    print("All cases have been processed.")

def createCSVResultFile(case_list_file, summary_file_name):
    def extract_key_results(case_file):
        try:
            with open(case_file, 'r') as res_file:
                res_lines = res_file.readlines()

            keyres1 = None
            keyres2 = None

            # Search for key results in the log file
            for line in res_lines:
                if re.search(r'Maximum BS curvature', line):
                    keyres1 = line.strip().split(':')[-1].strip()
                if re.search(r'Maximum curvature', line):
                    keyres2 = line.strip().split(':')[-1].strip()

            if keyres1 and keyres2:
                return {
                    "case_name": case_file.rstrip('.log'),
                    "maximum_bs_curvature": keyres1,
                    "maximum_curvature": keyres2
                }
            else:
                print(f"Key results not found in {case_file}")
                return None

        except FileNotFoundError:
            print(f"File not found: {case_file}")
            return None
        except Exception as e:
            print(f"Error processing {case_file}: {e}")
            return None

    try:
        with open(case_list_file, 'r') as f:
            cases = f.readlines()

        results = []
        for case in cases:
            case_file = case.strip().replace('.inp', '.log')
            result = extract_key_results(case_file)
            if result:
                results.append(result)

        with open(summary_file_name, 'w', newline='') as summary_file:
            writer = csv.DictWriter(summary_file, fieldnames=["case_name", "maximum_bs_curvature", "maximum_curvature"])
            writer.writeheader()
            writer.writerows(results)

    except FileNotFoundError:
        print(f"File not found: {case_list_file}")
    except Exception as e:
        print(f"Error processing {case_list_file}: {e}")

def save_results_to_excel(csv_file, excel_file, threshold):
    def generate_2d_array():
        # Load CSV data
        df = pd.read_csv(csv_file)

        # Initialize rows, columns, and data storage
        rows = []
        cols = []
        data = {}

        # Process results to extract rows, columns, and map data
        for _, result in df.iterrows():
            case_name = result['case_name']
            parts = case_name.split('-')

            if len(parts) >= 3:
                row = float(parts[1])  # First number
                col = float(parts[2])  # Second number

                # Add unique rows and columns
                if row not in rows:
                    rows.append(row)
                if col not in cols:
                    cols.append(col)

                # Store data in a dictionary with (row, col) as key
                data[(row, col)] = float(result['maximum_curvature'])

        # Sort rows and columns in ascending order
        rows.sort()
        cols.sort()

        # Generate 2D DataFrame
        table_data = {
            row: [data.get((row, col), None) for col in cols]
            for row in rows
        }
        df = pd.DataFrame.from_dict(table_data, orient='index', columns=cols)

        # Return rows, cols, data, and DataFrame
        return rows, cols, data, df

    def export_to_excel(df, file_name):
        # Save the DataFrame to an Excel file
        df.to_excel(file_name, sheet_name='Curvature Data', index_label='Row')

        # Apply conditional formatting
        wb = load_workbook(file_name)
        sheet = wb['Curvature Data']

        # Define color fills
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Loop through DataFrame values and apply formatting
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):  # Data starts at Excel row 2
            for col_idx, value in enumerate(row, start=2):  # Data starts at Excel column B
                if value is not None:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if value < threshold:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

        # Save the workbook
        wb.save(file_name)
        print(f"Data exported to {file_name} with conditional formatting.")

    # Generate the 2D array and print it
    rows, cols, data, df = generate_2d_array()
    print("Generated 2D Array:")
    print(df)

    # Export the 2D array to Excel
    export_to_excel(df, excel_file)


def main():
    #run_threads() #running the threads will create a lot of log files, which will be used to create the csv file.
    #createCSVResultFile('testbsengine-cases.txt', 'testbsengine-summary.csv') #creates the csv file that is used to create the excel file.
    save_results_to_excel('testbsengine-summary.csv', 'testbsengine-summary.xlsx', 0.0417592324286036) #creates the excel file with conditional formatting.

main()

#kan forbedres ved å finne en annen måte å starte tråder, fungerer for øyeblikket.