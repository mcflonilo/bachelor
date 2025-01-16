import os
import subprocess
from queue import Queue
import threading
import time
import re
import csv
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


class BSParams:
    def __init__(self, ID, SL, OD, CL, TL, TOD, MAT, MATID):
        self.ID = ID
        self.SL = SL
        self.OD = OD
        self.CL = CL
        self.TL = TL
        self.TOD = TOD
        self.MAT = MAT
        self.MATID = MATID

class RiserParams:
    def __init__(self, length, EA, EI, GT, m):
        self.length = length
        self.EA = EA
        self.EI = EI
        self.GT = GT
        self.m = m

def runBSEngine(case, case_queue):
        current_directory = os.getcwd()
        process = subprocess.Popen(f".\\bsengine -b {case}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=current_directory)
        stdout, stderr = process.communicate()
        print(f"stdout: {stdout}")
        print(f"stderr: {stderr}")
        if process.returncode != 0:
            print(f"Error encountered with case: {case}. Re-queuing the case.")
            case_queue.put(case)

def createCaseQueue():
    cases = open('testbsengine-cases.txt', 'r').readlines()
    cases = [case.strip() for case in cases]
    case_queue = Queue()
    for case in cases:
        case_queue.put(case)
    return case_queue

def run_threads():
    case_queue = Queue()
    case_queue = createCaseQueue()
    print(f"Number of cases: {case_queue.qsize()}.")
    max_threads = os.cpu_count() - 1
    print(f"Max threads: {max_threads}.")
    threads = []

    while not case_queue.empty():
        while threading.active_count() > max_threads:
            time.sleep(1)
        case = case_queue.get()
        print(f"Running case: {case}.")
        thread = threading.Thread(target=runBSEngine, args=(case,case_queue))
        time.sleep(1)
        thread.start()
        threads.append(thread)
        print(f"Active threads: {threading.active_count()}.")

    for thread in threads:
        thread.join()
    print("All cases have been processed.")

def createCSVResultFile(case_list_file, summary_file_name):
    def extract_key_results(case_file):
        try:
            with open(case_file, 'r') as res_file:
                res_lines = res_file.readlines()

            keyres1 = None
            keyres2 = None

            # Search for key results in the log file
            for line in res_lines:
                if re.search(r'Maximum BS curvature', line):
                    keyres1 = line.strip().split(':')[-1].strip()
                if re.search(r'Maximum curvature', line):
                    keyres2 = line.strip().split(':')[-1].strip()

            if keyres1 and keyres2:
                return {
                    "case_name": case_file.rstrip('.log'),
                    "maximum_bs_curvature": keyres1,
                    "maximum_curvature": keyres2
                }
            else:
                print(f"Key results not found in {case_file}")
                return None

        except FileNotFoundError:
            print(f"File not found: {case_file}")
            return None
        except Exception as e:
            print(f"Error processing {case_file}: {e}")
            return None

    try:
        with open(case_list_file, 'r') as f:
            cases = f.readlines()

        results = []
        for case in cases:
            case_file = case.strip().replace('.inp', '.log')
            result = extract_key_results(case_file)
            if result:
                results.append(result)

        with open(summary_file_name, 'w', newline='') as summary_file:
            writer = csv.DictWriter(summary_file, fieldnames=["case_name", "maximum_bs_curvature", "maximum_curvature"])
            writer.writeheader()
            writer.writerows(results)

    except FileNotFoundError:
        print(f"File not found: {case_list_file}")
    except Exception as e:
        print(f"Error processing {case_list_file}: {e}")

def save_results_to_excel(csv_file, excel_file, threshold):
    def generate_2d_array():
        # Load CSV data
        df = pd.read_csv(csv_file)

        # Initialize rows, columns, and data storage
        rows = []
        cols = []
        data = {}

        # Process results to extract rows, columns, and map data
        for _, result in df.iterrows():
            case_name = result['case_name']
            parts = case_name.split('-')

            if len(parts) >= 3:
                row = float(parts[1])  # First number
                col = float(parts[2])  # Second number

                # Add unique rows and columns
                if row not in rows:
                    rows.append(row)
                if col not in cols:
                    cols.append(col)

                # Store data in a dictionary with (row, col) as key
                data[(row, col)] = float(result['maximum_curvature'])

        # Sort rows and columns in ascending order
        rows.sort()
        cols.sort()

        # Generate 2D DataFrame
        table_data = {
            row: [data.get((row, col), None) for col in cols]
            for row in rows
        }
        df = pd.DataFrame.from_dict(table_data, orient='index', columns=cols)

        # Return rows, cols, data, and DataFrame
        return rows, cols, data, df

    def export_to_excel(df, file_name):
        # Save the DataFrame to an Excel file
        df.to_excel(file_name, sheet_name='Curvature Data', index_label='Row')

        # Apply conditional formatting
        wb = load_workbook(file_name)
        sheet = wb['Curvature Data']

        # Define color fills
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Loop through DataFrame values and apply formatting
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):  # Data starts at Excel row 2
            for col_idx, value in enumerate(row, start=2):  # Data starts at Excel column B
                if value is not None:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if value < threshold:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

        # Save the workbook
        wb.save(file_name)
        print(f"Data exported to {file_name} with conditional formatting.")

    # Generate the 2D array and print it
    rows, cols, data, df = generate_2d_array()
    print("Generated 2D Array:")
    print(df)

    # Export the 2D array to Excel
    export_to_excel(df, excel_file)

def generate_case_files(length, EA, EI, GT, m, cases, ID, SL, OD, CL, TL, TOD, MAT, MATID):
    inp1 = open('bsengine-cases.txt', 'w')

    x = 0 # Counter
    y = 0
    

    for i in cases:
        x = 0
        y = y + 1
        for j in OD:
            x = 0
            for k in CL:
                x = 0
                for l in MAT:
                    x = 0
                    q = MATID[x]
                    x = x + 1
                    # Create a directory for the case files if it doesn't exist
                    if not os.path.exists('case_files'):
                        os.makedirs('case_files')

                    inp = open(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp"), 'w')
                    inp1.write(os.path.join('case_files', "Case" + str(y) + '-' + str(j) + '-' + str(k) + '-' + str(q) + ".inp") + '\n')

                    inp.write('BEND STIFFENER DATA'+'\n')
                    inp.write("' ID   NSEG"+'\n')
                    inp.write("' inner diameter      Number of linear segments"+'\n')
                    inp.write(str(ID) + "   3"+'\n')
                    inp.write("' LENGTH   NEL   OD1    OD2  DENSITY LIN/NOLIN        EMOD/MAT-ID"+'\n')
                    inp.write(str(SL) + "  50  " + str(j) + "  " + str(j) + "  2000  LIN  10000.E06"+'\n')
                    inp.write(str(k) + "  100  " + str(j) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write(str(TL) + "  50  " + str(TOD) + "  " + str(TOD) + "  1150  " + str(l)+'\n')
                    inp.write("'-----------------------------------------------------------------------"+'\n')
                    inp.write("'"+'\n')
                    inp.write("RISER DATA"+'\n')
                    inp.write("'SRIS,  NEL   EI,    EA,      GT     Mass"+'\n')
                    inp.write("' (m)         kNm^2  kN              kg/m"+'\n')
                    inp.write(str(length) + "  100  " + str(EI) + "  " + str(EA) + "  " + str(GT) + "  " + str(m) +'\n')
                    inp.write("'"+'\n')
                    inp.write("' mandatory data group"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("ELEMENT PRINT"+'\n')
                    inp.write("'NSPEC"+'\n')
                    inp.write("3"+'\n')
                    inp.write("'IEL1    IEL2"+'\n')
                    inp.write("1         9"+'\n')
                    inp.write("10        30"+'\n')
                    inp.write("70        80"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE SYSTEM DATA TEST PRINT"+'\n')
                    inp.write("'IFSPRI 1/2"+'\n')
                    inp.write("1"+'\n')
                    inp.write("'2"+'\n')
                    inp.write("' -------------------------------------------------"+'\n')
                    inp.write("FE ANALYSIS PARAMETERS"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'  finite element analysis parameters"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'TOLNOR  MAXIT"+'\n')
                    inp.write("1.E-07  30"+'\n')
                    inp.write("'DSINC,DSMIN,DSMAX,"+'\n')
                    inp.write("0.01  0.001  0.1"+'\n')
                    inp.write("'3.0  0.01 10."+'\n')
                    inp.write("'5.  0.1 10."+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("CURVATURE RANGE"+'\n')
                    inp.write("'----------------------------------------------"+'\n')
                    inp.write("'CURMAX  - Maximum curvature"+'\n')
                    inp.write("'NCURV   - Number of curvature levels"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'CURMAX (1/m),NCURV"+'\n')
                    inp.write("'0.5       30"+'\n')
                    inp.write("0.2       100"+'\n')
                    inp.write("'---------------------------------------------------"+'\n')
                    inp.write("FORCE"+'\n')
                    inp.write("'Relang  tension"+'\n')
                    inp.write("'(deg)   (kN)"+'\n')
                    inp.write(str(i) + '\n')
                    inp.write("'8.00   400.0"+'\n')
                    inp.write("'16.5   500.0"+'\n')
                    inp.write("'19.0   550.0"+'\n')
                    inp.write("'19.1   600.0"+'\n')
                    inp.write("'18.6   650.0"+'\n')
                    inp.write("'17.5   700.0"+'\n')
                    inp.write("'14.0   775.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("'----------------------------------------------------"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_15"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.0 0.0"+'\n')
                    inp.write("0.005   1.40E+03"+'\n')
                    inp.write("0.010   2.57E+03"+'\n')
                    inp.write("0.015   3.61E+03"+'\n')
                    inp.write("0.020   4.55E+03"+'\n')
                    inp.write("0.025   5.36E+03"+'\n')
                    inp.write("0.030   6.03E+03"+'\n')
                    inp.write("0.035   6.59E+03"+'\n')
                    inp.write("0.040   7.02E+03"+'\n')
                    inp.write("0.045   7.37E+03"+'\n')
                    inp.write("0.050   7.67E+03"+'\n')
                    inp.write("0.055   7.92E+03"+'\n')
                    inp.write("0.060   8.13E+03"+'\n')
                    inp.write("0.065   8.31E+03"+'\n')
                    inp.write("0.070   8.47E+03"+'\n')
                    inp.write("0.075   8.61E+03"+'\n')
                    inp.write("0.080   8.74E+03"+'\n')
                    inp.write("0.085   8.86E+03"+'\n')
                    inp.write("0.090   8.96E+03"+'\n')
                    inp.write("0.095   9.06E+03"+'\n')
                    inp.write("0.100   9.10E+03"+'\n')
                    inp.write("'"+'\n')
                    inp.write("MATERIAL DATA"+'\n')
                    inp.write("' Material identifier"+'\n')
                    inp.write("60D_30"+'\n')
                    inp.write("'NMAT - Number of points in stress/strain table for BS material"+'\n')
                    inp.write("21"+'\n')
                    inp.write("' strain   stress (kPa)    - Nmat input lines"+'\n')
                    inp.write("0.000   0.0"+'\n')
                    inp.write("0.005   1100.0"+'\n')
                    inp.write("0.010   2060.0"+'\n')
                    inp.write("0.015   2910.0"+'\n')
                    inp.write("0.020   3690.0"+'\n')
                    inp.write("0.025   4370.0"+'\n')
                    inp.write("0.030   4950.0"+'\n')
                    inp.write("0.035   5420.0"+'\n')
                    inp.write("0.040   5810.0"+'\n')
                    inp.write("0.045   6120.0"+'\n')
                    inp.write("0.050   6400.0"+'\n')
                    inp.write("0.055   6640.0"+'\n')
                    inp.write("0.060   6840.0"+'\n')
                    inp.write("0.065   7030.0"+'\n')
                    inp.write("0.070   7180.0"+'\n')
                    inp.write("0.075   7330.0"+'\n')
                    inp.write("0.080   7470.0"+'\n')
                    inp.write("0.085   7590.0"+'\n')
                    inp.write("0.090   7710.0"+'\n')
                    inp.write("0.095   7810.0"+'\n')
                    inp.write("0.100   7920.0"+'\n')
                    inp.write("'"+'\n')
                    inp.write("'EXPORT MATERIAL DATA"+'\n')
                    inp.write("'--------------------------------------------------"+'\n')
                    inp.write("' IMEX   = 1 : tabular  =2 riflex format"+'\n')
                    inp.write("'  1"+'\n')
                    inp.write("'---------------------------------------------"+'\n')
                    inp.write("end"+'\n')
                    inp.write("'mandatory data group"+'\n')
                    inp.write("'---------------------------------------------"+'\n')

    inp.close()
    inp1.close()

def main():
    BSParams1 = BSParams(1, 0.5, 0.2, 0.1, 0.05, 0.01, [60, 60], ['60D_15', '60D_30'])
    RiserParams1 = RiserParams(100, 100, 100, 100, 100)
    #generate_case_files(RiserParams1.length, RiserParams1.EA, RiserParams1.EI, RiserParams1.GT, RiserParams1.m, [8, 16.5, 19, 19.1, 18.6, 17.5, 14], BSParams1.ID, BSParams1.SL, BSParams1.OD, BSParams1.CL, BSParams1.TL, BSParams1.TOD, BSParams1.MAT, BSParams1.MATID)
    run_threads() #running the threads will create a lot of log files, which will be used to create the csv file.
    createCSVResultFile('testbsengine-cases.txt', 'testbsengine-summary.csv') #creates the csv file that is used to create the excel file.
    save_results_to_excel('testbsengine-summary.csv', 'testbsengine-summary.xlsx', 0.0417592324286036) #creates the excel file with conditional formatting.

main()

#kan forbedres ved å finne en annen måte å starte tråder, fungerer for øyeblikket.